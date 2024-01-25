from flask import Flask, render_template
from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import Workbook, load_workbook
import os


app = Flask(__name__, static_folder='static')
excel_file_path = 'TelegramBot/Donation.xlsx'
from flask import Flask, render_template
from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__, static_folder='static')
excel_file_path = os.path.join(os.path.dirname(__file__), 'donations.xlsx')

# Ensure the Excel file exists and has the correct structure
if not os.path.exists(excel_file_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value=0)  # Initialize total donations to 0
    workbook.save(excel_file_path)

def update_donations():
    global total_donations
    try:
        if os.path.exists(excel_file_path):
            workbook = load_workbook(excel_file_path)
            sheet = workbook.active
            total_donations = sheet.cell(row=1, column=1).value
            total_donations += 100  # Increment by a random value (replace with your logic)
            sheet.cell(row=1, column=1, value=total_donations)
            workbook.save(excel_file_path)
    except Exception as e:
        print(f"Error updating donations: {e}")

# Schedule the background task to run every 5 hours
scheduler = BackgroundScheduler()
scheduler.add_job(update_donations, 'interval', hours=5)
scheduler.start()

@app.route('/deposit/<deposit_id>')
def deposit(deposit_id):
    return render_template('index2.html', deposit_id=deposit_id)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)
excel_file_path = 'donations.xlsx'

# Ensure the Excel file exists and has the correct structure
if not os.path.exists(excel_file_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value=0)  # Initialize total donations to 0
    workbook.save(excel_file_path)

def update_donations():
    global total_donations
    try:
        if os.path.exists(excel_file_path):
            workbook = load_workbook(excel_file_path)
            sheet = workbook.active
            total_donations = sheet.cell(row=1, column=1).value
            total_donations += 100  # Increment by a random value (replace with your logic)
            sheet.cell(row=1, column=1, value=total_donations)
            workbook.save(excel_file_path)
    except Exception as e:
        print(f"Error updating donations: {e}")


# Schedule the background task to run every 5 hours
scheduler = BackgroundScheduler()
scheduler.add_job(update_donations, 'interval', hours=5)
scheduler.start()


@app.route('/deposit/<deposit_id>')
def deposit(deposit_id):
    return render_template('index2.html', deposit_id=deposit_id)
